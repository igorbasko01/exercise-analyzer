import csv

from openpyxl import load_workbook

from data_handler import extract_exercise_names, explode_exercises

original_exercise_names = ['Chest', 'Squat', 'Row', 'Biceps W', 'Pullups']
new_exercise_names = ['Bench Press', 'Squat', 'Bent Over Row', 'EZ Barbell Curl', 'Pullups']

wb = load_workbook('Exercise.xlsx', data_only=False)

ws = wb['Combined']

# Extract the first row, converts tuple to list, the next is used to actually extract the row from the generator.
first_row = list(map(lambda x: x.value, list(next(ws.iter_rows(min_row=1, max_row=1)))))

exercise_names = extract_exercise_names(first_row, original_exercise_names)
print('Exercise names:', exercise_names)

if not exercise_names:
    print('No exercise names found')
    exit()

original_headers = ['Date'] + exercise_names

whole_data = []
for row in ws.iter_rows(min_row=2):
    row_data = []
    for cell in row:
        if cell.value is not None:
            row_data.append(cell.value)
        else:
            row_data.append(None)
    whole_data.append(row_data)

print('Whole data:', whole_data)

exploded = explode_exercises(whole_data,
                             original_headers,
                             alternative_exercise_names=new_exercise_names)

new_headers = ['Date', 'Exercise', 'Weight', 'Reps']
exploded_with_headers = [new_headers] + list(exploded)
print('Exploded:', exploded_with_headers)

with open('output.csv', 'w') as f:
    writer = csv.writer(f)
    writer.writerows(exploded_with_headers)